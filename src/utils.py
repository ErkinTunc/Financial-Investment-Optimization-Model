import openpyxl


def lecture_donnees(file_path='data-folder/data-exemple/exemple.xlsx'):
    """Read problem data from an Excel file.

    Parameters:
    - file_path (str): path to the Excel file

    Returns:
    - n (int): total number of time periods
    - tau0 (float): base interest rate per period (as a decimal)
    - placements (list[tuple[float,int,int]]): (tau_k, d_k, f_k)
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    n = int(sheet.cell(row=1, column=1).value)
    tau0 = float(sheet.cell(row=2, column=1).value) / 100.0  # percent -> decimal

    placements = []
    row = 3
    while True:
        tauk = sheet.cell(row=row, column=1).value
        dk = sheet.cell(row=row, column=2).value
        fk = sheet.cell(row=row, column=3).value

        if tauk is None or dk is None or fk is None:
            break

        placements.append((float(tauk) / 100.0, int(dk), int(fk)))
        row += 1

    return n, tau0, placements


def index_placements_by_end_time(n, placements):
    """Group products by their end time f_k.

    This turns the DP from O(n·m) (scan all products at each t)
    into O(n+m) (scan only products ending at t).

    Returns:
    - ending_at (list[list[tuple[float,int]]]): ending_at[t] = [(tau_k, d_k), ...] for products with f_k == t
    """
    ending_at = [[] for _ in range(n + 1)]
    for tau_k, d_k, f_k in placements:
        if 0 <= f_k <= n:
            ending_at[f_k].append((tau_k, d_k))
    return ending_at


def optimiz_coef_indexed(t, coef, tau0, ending_at, prev):
    """Compute optimal Coef(t) using pre-indexed products.

    Parameters:
    - t (int): current time step
    - coef (list[float]): already computed Coef values up to t-1
    - tau0 (float): base interest rate per period
    - ending_at (list[list[tuple[float,int]]]): products indexed by end time
    - prev (list[int|None]): predecessor time index chosen for each t

    Returns:
    - best (float): optimal Coef(t)
    """
    # Option 1: base interest from (t-1) -> t
    best = coef[t - 1] * (1.0 + tau0)
    prev[t] = t - 1

    # Option 2: only products that end at time t
    for tau_k, d_k in ending_at[t]:
        val = coef[d_k] * (1.0 + tau_k)
        if val > best:
            best = val
            prev[t] = d_k

    return best


def reconstruct_path(prev, end):
    """Reconstruct the optimal investment path from a predecessor array.

    Parameters:
    - prev (list[int|None]): prev[t] stores the previous time chosen to reach t optimally
    - end (int): final time step (e.g., n)

    Returns:
    - path (list[tuple[int,int]]): list of (from, to) steps from 0 to end
    """
    path = []
    t = end

    while t is not None and prev[t] is not None:
        p = prev[t]
        path.append((p, t))
        t = p

    path.reverse()
    return path


def enumerate_paths(n, tau0, placements):
    """Enumerate all possible paths from 0 to n (debug / sanity checks).

    Warning: exponential in the worst case. Use only for small instances.

    Parameters:
    - n (int): total number of time periods
    - tau0 (float): base interest rate per period (e.g. 0.009)
    - placements (list[tuple[float,int,int]]): (tau_k, d_k, f_k)
    """
    from collections import defaultdict

    graph = defaultdict(list)

    for t in range(n):
        graph[t].append((t + 1, 1.0 + tau0))

    for tau_k, d_k, f_k in placements:
        graph[d_k].append((f_k, 1.0 + tau_k))

    all_paths = []

    def explore_paths(t, path, coef_mult):
        if t == n:
            all_paths.append((path[:], coef_mult))
            return

        for next_t, multiplier in graph[t]:
            if next_t <= n:
                path.append((t, next_t))
                explore_paths(next_t, path, coef_mult * multiplier)
                path.pop()

    explore_paths(0, [], 1.0)

    for i, (path, coef_mult) in enumerate(all_paths, 1):
        print(f"Path {i}: {path} -> Total Coefficient: {coef_mult:.6f}")
